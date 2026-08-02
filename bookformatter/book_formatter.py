"""
book_formatter.py
=================

Reads every scraped book .xlsx in a folder and writes a NEW, slim .xlsx per file
holding the finished listing content, keyed by the product name so the two files
can be joined with VLOOKUP:

    A  Name              -> product name copied VERBATIM from the input file
    B  Revised Name      -> title in English
    C  Revised Author    -> author in English
    D  Book Details      -> HTML bullet list (Author, Publisher, ISBN, ...)
    E  Book Description  -> one SEO-friendly paragraph (90-150 words)
    F  Meta Description  -> <= 160 char meta description starting "Buy Online"
    G  Listing HTML      -> book details + description, ready to paste
    H  Tags              -> 5-12 comma-separated Shopify collection tags

Column A is deliberately an untouched copy of the source cell - trailing spaces
included - because VLOOKUP needs an exact match, and it sits first because VLOOKUP
only searches the leftmost column of its range:

    =VLOOKUP($A2, [output.xlsx]Books!$A:$H, 7, FALSE)   -> Listing HTML

No other scraped column is copied. If you need one back for cross-checking (e.g.
the product URL), add it with --carry link.

The two source files have DIFFERENT schemas (different headers, different
languages, different junk columns), so nothing is positional: headers are matched
against alias sets, values are normalised, and whatever is left over is still
handed to the model as "additional data". Adding a third scraper usually needs
nothing more than one extra alias.

Text that is ALREADY English is never sent for translation: a Latin-script title
or author is copied straight through and left out of the model's output schema,
so no tokens are spent re-writing English into English.

Usage
-----
    pip install -r requirements.txt
    setx OPENAI_API_KEY "sk-..."          (or put it in a .env file)

    python book_formatter.py                       # process every .xlsx in this folder
    python book_formatter.py --files book1.xlsx    # only one file
    python book_formatter.py --limit 5             # smoke-test on 5 rows per file
    python book_formatter.py --workers 8 --model gpt-4o
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from html import escape
from pathlib import Path
from types import SimpleNamespace

import openpyxl
from openpyxl.styles import Alignment, Font
from openai import OpenAI

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:  # dotenv is optional
    pass


# --------------------------------------------------------------------------- #
# 0. OUTPUT - every message goes through log() so the GUI can capture it
# --------------------------------------------------------------------------- #
def log(message: str = "") -> None:
    """Emit one line of progress. Replaced by the GUI with a queue writer."""
    _SINK(str(message))


def set_log_sink(sink) -> None:
    global _SINK
    _SINK = sink


_SINK = print


# --------------------------------------------------------------------------- #
# 1. SCHEMA MAPPING - how any scraper's headers map onto our canonical fields
# --------------------------------------------------------------------------- #
# Aliases are listed in priority order: the first alias that holds a non-empty
# value wins. That is how "Author" beats the duplicate "Writer" column in the
# Baatighar file, while a row that only filled "Writer" still resolves.
COLUMN_ALIASES: dict[str, list[str]] = {
    "name":            ["name", "book name", "title", "book title", "product name", "product"],
    "author":          ["author", "authors", "author name", "writer", "writers", "by"],
    "translator":      ["translator", "translated by"],
    "editor":          ["editor", "edited by"],
    "publisher":       ["publisher", "published by", "imprint", "publication"],
    "isbn":            ["isbn", "isbn13", "isbn 13", "isbn10", "isbn 10"],
    "language":        ["language", "languages", "lang"],
    "country":         ["country", "country of origin", "origin"],
    "binding":         ["binding", "format", "cover", "cover type", "book format", "binding type"],
    "edition":         ["edition", "edition number"],
    "first_published": ["first published", "publication year", "year of publication",
                        "published on", "publish date", "year"],
    "pages":           ["pages", "page", "no of pages", "number of pages", "page count"],
    "dimensions":      ["dimensions", "dimension", "size", "book size"],
    "weight":          ["weight", "item weight"],
    "category":        ["category", "categories", "genre", "genres", "subject", "subjects", "tags"],
    "source_blurb":    ["description", "book description", "summary", "about the book",
                        "overview", "details", "synopsis"],
    "link":            ["link", "url", "product url", "product link", "source"],
}

# Commerce / scraping noise. Never shown to the model - the prompts explicitly
# forbid price, stock and delivery information from appearing in the output.
IGNORED_COLUMNS = {
    "sellingprice", "originalprice", "price", "mrp", "discount", "discountpercentage",
    "otherfees", "fees", "shipping", "delivery", "availability", "stock", "stockstatus",
    "sku", "productcode", "barcode", "image", "imageurl", "thumbnail", "rating",
    "reviews", "quantity", "instock",
}

# Only consulted when --infer-country is passed. The prompts say "never invent
# missing information", so vendor-country inference is opt-in, not the default.
DOMAIN_COUNTRY = {
    "baatighar.com": "Bangladesh",
    "exoticindiaart.com": "India",
    "exoticindia.com": "India",
}

# Field order used when rendering the Book Details bullet list.
DETAIL_ORDER = [
    ("author",          "Author"),
    ("translator",      "Translator"),
    ("editor",          "Editor"),
    ("publisher",       "Publisher"),
    ("isbn",            "ISBN"),
    ("language",        "Language"),
    ("country",         "Country"),
    ("binding",         "Binding"),
    ("edition",         "Edition"),
    ("first_published", "First Published"),
    ("pages",           "Pages"),
    ("dimensions",      "Dimensions"),
    ("weight",          "Weight"),
]

# Column A is the VLOOKUP key: the product name copied verbatim from the input
# file - not cleaned, not trimmed - so that a lookup against the source sheet
# matches exactly. VLOOKUP only searches the first column of its range, which is
# why the key sits at A rather than at the end.
KEY_COLUMN = "Name"

# B..H. The listing HTML deliberately excludes the meta description - that text is
# taken from column F, which is why there is no meta-HTML column.
OUTPUT_COLUMNS = [
    "Revised Name",             # B
    "Revised Author",           # C
    "Book Details",             # D
    "Book Description",         # E
    "Meta Description",         # F
    "Listing HTML",             # G - book details + book description only
    "Tags",                     # H - comma-separated Shopify collection tags
]

# Source columns carried into the output file. Empty by default: the new file holds
# only generated content. Opt back in per run with e.g. --carry link,isbn.
DEFAULT_CARRY: list[str] = []

CARRY_HEADERS = {
    "name": "Name",
    "author": "Author",
    "link": "Link",
    "isbn": "ISBN",
    "publisher": "Publisher",
    "language": "Language",
    "image": "Image",
}

# Column widths in the output sheet, by header name.
COLUMN_WIDTHS = {
    "Name": 40, "Author": 24, "Link": 40, "ISBN": 16, "Publisher": 24,
    "Language": 12, "Image": 40,
    "Revised Name": 38, "Revised Author": 24, "Book Details": 55,
    "Book Description": 70, "Meta Description": 55,
    "Listing HTML": 70, "Tags": 50,
}

# ---------------------------------------------------------------------------
# Per-sheet rules. Matched against the source filename first, then the product
# URL, so a rule only ever affects the sheet it was written for.
#
#   bilingual_title -- render the title as "English (Original Script)", e.g.
#                      "Swapna-Bishleshon (স্বপ্ন-বিশ্লেষণ)". Applied only to rows
#                      whose title really is non-Latin; English titles stay plain.
# ---------------------------------------------------------------------------
SHEET_PROFILES = [
    {"match": ["baatighar"],               "bilingual_title": True},
    {"match": ["exoticindia", "exotic"],   "bilingual_title": False},
]

DEFAULT_PROFILE = {"match": [], "bilingual_title": False}


def get_profile(path: Path, sample_link: str = "") -> dict:
    """Pick the rule set for one workbook (filename wins, product URL is fallback)."""
    haystack = f"{path.stem} {sample_link}".lower()
    for profile in SHEET_PROFILES:
        if any(token in haystack for token in profile["match"]):
            return profile
    return DEFAULT_PROFILE

EXCEL_CELL_LIMIT = 32_000  # real limit is 32767; leave headroom


# --------------------------------------------------------------------------- #
# 2. VALUE CLEANING - per-field normalisers that absorb the file differences
# --------------------------------------------------------------------------- #
def _norm_header(value) -> str:
    """Lower-case, strip punctuation/whitespace so 'ISBN-13' == 'isbn 13'."""
    text = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())
    return re.sub(r"\s+", " ", text).strip()


def _clean(value) -> str:
    """Excel cell -> tidy string (NBSP, zero-width chars, stray whitespace)."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("​", "")
    text = re.sub(r"\s+", " ", text).strip()
    return "" if text.lower() in {"n/a", "na", "none", "-", "--", "null"} else text


def _clean_language(value: str) -> str:
    """'Bengali / বাংলা' -> 'Bengali'; 'English Text' -> 'English'; 'KANNADA' -> 'Kannada'."""
    text = re.split(r"[/|,;]", value)[0]
    text = re.sub(r"\b(text|language|edition)\b", "", text, flags=re.I).strip(" -–—")
    return text.title() if text.isupper() or text.islower() else text


def _clean_weight(value: str) -> str:
    """'Weight 410 gm' -> '410 gm'."""
    return re.sub(r"^\s*weight\s*[:\-]?\s*", "", value, flags=re.I).strip()


def _clean_pages(value: str) -> str:
    """'200 pages' / '200.0' -> '200'."""
    match = re.search(r"\d[\d,]*", value)
    return match.group(0).replace(",", "") if match else ""


def _clean_isbn(value: str) -> str:
    """Drop separators; reject the fake ISBNs some scrapers emit for un-ISBN'd items."""
    digits = re.sub(r"[^0-9Xx]", "", value)
    if len(digits) not in (10, 13):
        return ""
    # e.g. Baatighar writes '1186000000000' (an internal id) when there is no ISBN.
    if re.fullmatch(r"\d{4}0{6,}", digits):
        return ""
    return digits.upper()


def _clean_binding(value: str) -> str:
    """'PAPERBACK' -> 'Paperback'."""
    return value.title() if value.isupper() else value


# Any character outside Latin/punctuation/digits means the text is not English yet.
NON_LATIN_RE = re.compile(
    r"[^\x09\x0a\x0d\x20-\x7e\u00a0-\u024f\u1e00-\u1eff"
    r"\u2000-\u206f\u20a0-\u20cf\u2100-\u214f]"
)


def is_english(text: str) -> bool:
    """True when the text is already Latin script and can be copied verbatim.

    'Sigmund Freud' -> True (no API call needed)
    'স্বপ্ন-বিশ্লেষণ' -> False
    'ಶ್ರೀವಿದ್ಯಾ- Srividyamanyatirtha (Kannada)' -> False (mixed; the model picks the
    English form that is already embedded in it)
    """
    return bool(text) and not NON_LATIN_RE.search(text)


FIELD_CLEANERS = {
    "language": _clean_language,
    "weight": _clean_weight,
    "pages": _clean_pages,
    "isbn": _clean_isbn,
    "binding": _clean_binding,
}


def build_header_map(header_row: tuple) -> tuple[dict[str, int], dict[int, str]]:
    """Return (canonical_field -> column index, leftover column index -> header)."""
    headers = {idx: _norm_header(cell) for idx, cell in enumerate(header_row)}
    field_to_cols: dict[str, list[int]] = {}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:  # priority order matters
            for idx, header in headers.items():
                if header == alias:
                    field_to_cols.setdefault(field, []).append(idx)

    claimed = {idx for cols in field_to_cols.values() for idx in cols}
    leftovers = {
        idx: str(header_row[idx]).strip()
        for idx, header in headers.items()
        if header and idx not in claimed and header.replace(" ", "") not in IGNORED_COLUMNS
    }
    return field_to_cols, leftovers


def extract_book(row: tuple, field_to_cols: dict, leftovers: dict, infer_country: bool) -> dict:
    """Row -> canonical dict. Unknown-but-useful columns land in 'additional'."""
    book: dict[str, str] = {}

    for field, columns in field_to_cols.items():
        for idx in columns:  # first alias with a real value wins
            raw = row[idx] if idx < len(row) else None
            value = _clean(raw)
            if value:
                book[field] = FIELD_CLEANERS.get(field, lambda v: v)(value)
                # The title is also kept byte-for-byte, trailing spaces and all:
                # it is the VLOOKUP key and must match the input cell exactly.
                if field == "name":
                    book["name_raw"] = str(raw)
                break
        book.setdefault(field, "")
    book.setdefault("name_raw", book.get("name", ""))

    # A "Description" column holding only the shop name ("Baatighar") is noise,
    # not a blurb. Anything under 40 chars is treated as unusable.
    if len(book.get("source_blurb", "")) < 40:
        book["source_blurb"] = ""

    if infer_country and not book.get("country"):
        link = book.get("link", "").lower()
        for domain, country in DOMAIN_COUNTRY.items():
            if domain in link:
                book["country"] = country
                break

    book["additional"] = {
        name: _clean(row[idx])
        for idx, name in leftovers.items()
        if idx < len(row) and _clean(row[idx])
    }
    return book


# --------------------------------------------------------------------------- #
# 3. PROMPTS - transcribed from the two .docx briefs in this folder
# --------------------------------------------------------------------------- #
DETAILS_SYSTEM_PROMPT = """You are an expert eCommerce content writer for an online bookstore.
Convert the raw scraped book information into a professional product listing.

RULES - BOOK DETAILS
- Include ONLY fields that are actually present in the supplied data. Never invent
  missing information and never output a field with an empty or guessed value.
- Never include SKU, Product Code, MRP, Price, Discount, Availability, Delivery
  information, Stock Status, Barcode or image URLs.
- Transliterate author, translator, editor and publisher names written in Bengali,
  Hindi, Kannada, Urdu or any other script into proper English romanisation.
  Use the established English spelling when the person or press is known in English.
- Keep Language as a plain name only: "Bengali", "English", "Kannada"
  (drop "/ বাংলা", "/ English", "Text", etc.).
- Normalise Edition to "1st Edition", "2nd Edition", "3rd Edition", ... A bare year
  in the edition field is an edition year, so output it as e.g. "2026 Edition" only
  if that is clearly what the source means; otherwise leave Edition out.
- Pages must be a plain number. Weight and Dimensions stay in their source units.

RULES - VALUES THAT ARE ALREADY ENGLISH
- Any value already written in Latin script is already English: copy it verbatim.
  Do not re-word it, re-order it, re-spell it, translate it or "improve" it.
  Only fix obvious ALL-CAPS by converting to normal title case.

RULES - REVISED NAME AND AUTHOR (only when requested below)
- revised_name: the book title written in English. Transliterate non-Latin scripts
  (Bengali, Devanagari, Kannada, Arabic ...). If the source title already contains an
  English form or a parenthetical romanisation, use that form as-is. Keep any
  meaningful qualifier such as a language or issue marker. Do not add words that are
  not implied by the original title.
- revised_author: the author name(s) in English, transliterated the same way and
  separated by ", ".

RULES - BOOK DESCRIPTION
- One single SEO-friendly paragraph of roughly 90-150 words. No bullet points,
  no headings, no line breaks.
- Natural, professional English that reads like publisher copy.
- Identify the genre from the supplied category/title/subject (Novel, Short Stories,
  Poetry, Horror, Children's Literature, Religion, Philosophy, Psychology, Self-Help,
  History, Politics, Economics, Science, Art, Archaeology ...).
- Describe what a reader can expect, grounded in the title, genre, author or subject.
  Never fabricate plot points, awards, sales figures or quotes.
- If it is a translated edition, mention the translator naturally.
- Never mention price, discount, delivery, SKU, availability, dimensions or page count.
- Avoid repetitive marketing filler.

OUTPUT
Return ONLY a JSON object containing exactly the keys listed under "return_keys" in
the user message - no more, no fewer. Possible keys:
  "revised_name"   -> string
  "revised_author" -> string
  "details"        -> object, keys drawn from: Author, Translator, Editor, Publisher,
                      ISBN, Language, Country, Binding, Edition, First Published,
                      Pages, Dimensions, Weight - omitting every unavailable field
  "description"    -> string
Keys that are not requested are already correct in the source data: leave them out
entirely rather than repeating them back."""

META_SYSTEM_PROMPT = """You are an expert SEO copywriter for an online bookstore.
Write ONE meta description for a book product page, optimised for click-through rate.

REQUIREMENTS
- English only. Maximum 160 characters including spaces; aim for 120-160.
- It MUST start with the words: Buy Online
- Naturally include the book title, and the author's name when one is supplied.
- Include the ISBN when one is supplied, written as "ISBN: 9789843915085".
- Mention the primary keyword/genre (Bengali Novel, Horror, Poetry, Psychology,
  Self-Help, History, Children's Book, Art, Religion ...) using only keywords that
  genuinely fit this book.
- Compelling and unique. No keyword stuffing, no needless repetition.
- No emojis, no quotation marks, no pricing, discounts, offers, shipping,
  availability or promotional claims. Never invent facts.

Return only the meta description text - no labels, no quotes, no extra commentary.

HARD CONSTRAINTS - these are checked programmatically, a violation is rejected:
- 160 characters maximum. Count them. A shorter description is always better than
  one that overruns.
- Quote the ISBN ONLY if an "ISBN" value is supplied in the user message, and then
  only that exact number. If no ISBN is supplied, the words "ISBN" must not appear
  at all. Never reuse an ISBN from an example - example numbers are illustrations of
  formatting only and belong to a different book.

Example of the FORM (its numbers belong to another book - never copy them):
Buy Online <Title> by <Author>, <short genre phrase>. ISBN: <the supplied ISBN>."""


TAGS_SYSTEM_PROMPT = """You are an expert librarian, bookseller, and book cataloguing specialist.
Your task is to assign CATEGORY TAGS to books for an ecommerce bookstore.

IMPORTANT
These tags are NOT SEO keywords.
These tags will be used to automatically group books into Shopify Collections.
Therefore, generate ONLY high-level and meaningful category tags.
DO NOT generate search phrases.
DO NOT generate promotional words.
DO NOT generate ISBN.
DO NOT generate publisher names.
DO NOT generate author names unless the author is itself a browsing category
(for example: Shakespeare, Rabindranath Tagore, Sigmund Freud, Stephen King).

Your objective is to answer:
"If a customer is browsing categories, under which collections should this book appear?"

RULES
1. Generate between 5 and 12 tags only.
2. Every tag must represent a category or collection.
3. Prefer broad categories first, followed by more specific ones.
4. Never repeat tags.
5. Never invent information.
6. Return ONLY comma-separated tags.
7. Use Title Case for English tags (leave established acronyms such as UPSC, NEET,
   SSC, UGC NET in capitals).
8. Keep tags short (1-3 words whenever possible).
9. Do not include explanations.

Generate tags from these types only.
- Main Subject: Psychology, History, Religion, Philosophy, Politics, Economics,
  Science, Medicine, Law, Education, Engineering, Technology, Computer Science
- Sub Subject: Psychoanalysis, Dream Analysis, Behavioural Science, Astrology,
  Vedic Astrology, Palmistry, Numerology, Yoga, Meditation, Ayurveda, Vedanta,
  Buddhism, Hinduism, Islam, Christianity, Sociology, Anthropology
- Literary Type: Novel, Fiction, Non Fiction, Poetry, Drama, Short Stories, Essays,
  Biography, Memoir, Autobiography, Research, Reference, Academic, Children
- Academic Stream: Commerce, Arts, Humanities, Medical, Engineering, Law
- Exam Category (if applicable): UPSC, SSC, WBPSC, Railway, NEET, JEE, UGC NET
- Language: Bengali Books, English Books, Hindi Books, Tamil Books, Malayalam Books
- Age Group (if applicable): Children, Young Adult, Adult
- Religious Category (if applicable): Hinduism, Buddhism, Islam, Christianity,
  Jainism, Sikhism
- Special Collections: Classics, Rare Books, Collected Works, Illustrated Books,
  Graphic Novels

Input will be provided in JSON.
Return ONLY comma-separated tags."""

TAG_MIN, TAG_MAX = 5, 12


def clean_tags(raw: str, book: dict) -> str:
    """Normalise the model's tag list and enforce the rules that code can enforce.

    Dedupes case-insensitively, drops ISBN-like and publisher-name tags, and caps
    the list at 12. Acronyms (UPSC, NEET) keep their capitals - blanket title-casing
    would turn them into 'Upsc'.
    """
    publisher = book.get("publisher", "").lower()
    seen: set[str] = set()
    tags: list[str] = []

    # The model sometimes prefixes the list with a label despite being told not to.
    raw = re.sub(r"^\s*(tags|category tags|output)\s*[:\-]\s*", "", raw or "", flags=re.I)

    for part in re.split(r"[,\n;]+", raw):
        tag = _clean(part).strip(" .-•*")
        if not tag or len(tag) > 40:
            continue
        if re.search(r"\d{5,}", tag):                       # an ISBN slipped through
            continue
        if publisher and tag.lower() in publisher:          # publisher name as a tag
            continue
        if tag.islower():                                   # leave UPSC / UGC NET alone
            tag = tag.title()
        key = tag.lower()
        if key in seen:
            continue
        seen.add(key)
        tags.append(tag)

    # The language collection is grounded in the source data, so add it if the model
    # forgot it. Nothing else is ever added - inventing categories is not allowed.
    language = book.get("language", "")
    if language:
        language_tag = f"{language} Books"
        if language_tag.lower() not in seen and len(tags) < TAG_MAX:
            tags.append(language_tag)

    return ", ".join(tags[:TAG_MAX])


META_LIMIT = 160
ISBN_MENTION_RE = re.compile(r"[,;\s]*\bISBNS?:?\s*([0-9][0-9\- ]{7,20}[0-9Xx])\.?", re.I)


def strip_wrong_isbn(meta: str, isbn: str) -> str:
    """Remove any ISBN that is not the book's own.

    The model has been observed copying the ISBN out of the prompt's example onto
    books that have none at all, so this is enforced in code rather than trusted
    to the prompt.
    """
    def keep_or_drop(match: re.Match) -> str:
        found = re.sub(r"[^0-9Xx]", "", match.group(1)).upper()
        return match.group(0) if isbn and found == isbn else ""

    cleaned = ISBN_MENTION_RE.sub(keep_or_drop, meta)
    cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" ,;-")
    if cleaned and cleaned[-1] not in ".!?":
        cleaned += "."
    return cleaned


def trim_meta(meta: str, limit: int = META_LIMIT) -> str:
    """Enforce the character cap, dropping whole trailing sentences where possible."""
    if len(meta) <= limit:
        return meta
    sentences = re.split(r"(?<=[.!?])\s+", meta)
    kept = ""
    for sentence in sentences:
        candidate = f"{kept} {sentence}".strip()
        if len(candidate) > limit:
            break
        kept = candidate
    if kept:
        return kept
    return meta[:limit].rsplit(" ", 1)[0].rstrip(" ,;-") + "."


def plan_translation(book: dict) -> dict:
    """Decide locally what actually needs the model.

    A Latin-script title or author is already English, so it is copied straight
    into the output and dropped from the model's return schema - no tokens burned
    translating English into English.
    """
    name, author = book.get("name", ""), book.get("author", "")
    return {
        "name_ok": is_english(name),
        "author_ok": is_english(author) or not author,
        "preset_name": name if is_english(name) else "",
        "preset_author": author if is_english(author) else "",
    }


def build_details_payload(book: dict, plan: dict) -> str:
    """Only real, cleaned values reach the model - blanks would invite invention."""
    data = {label: book[field] for field, label in DETAIL_ORDER if book.get(field)}

    title_label = "Title" if plan["name_ok"] else "Title (original, non-English)"
    data[title_label] = book.get("name", "")
    if book.get("category"):
        data["Category"] = book["category"]
    if book.get("source_blurb"):
        data["Source blurb"] = book["source_blurb"][:1500]
    for key, value in book.get("additional", {}).items():
        data.setdefault(key, value[:300])

    return_keys = ["details", "description"]
    if not plan["name_ok"]:
        return_keys.insert(0, "revised_name")
    if not plan["author_ok"]:
        return_keys.insert(0, "revised_author")

    return json.dumps(
        {"book_data": data, "return_keys": return_keys}, ensure_ascii=False, indent=2
    )


def build_meta_payload(book: dict, stage_one: dict) -> str:
    payload = {
        "Book Title (English)": stage_one.get("revised_name") or book.get("name", ""),
        "Author (English)": stage_one.get("revised_author", ""),
        "ISBN": book.get("isbn", ""),
        "Language": book.get("language", ""),
        "Category": book.get("category", ""),
        "Book Description": stage_one.get("description", "")[:900],
    }
    return json.dumps({k: v for k, v in payload.items() if v}, ensure_ascii=False, indent=2)


def build_tags_payload(book: dict, stage_one: dict) -> str:
    """The generated description carries the genre signal, so it is included.

    ISBN and publisher are deliberately left out: they must never become tags, and
    the surest way to stop that is not to show them.
    """
    payload = {
        "Book Title (English)": stage_one.get("revised_name") or book.get("name", ""),
        "Author (English)": stage_one.get("revised_author", ""),
        "Language": book.get("language", ""),
        "Category": book.get("category", ""),
        "Binding": book.get("binding", ""),
        "Book Description": stage_one.get("description", "")[:900],
    }
    return json.dumps({k: v for k, v in payload.items() if v}, ensure_ascii=False, indent=2)


# --------------------------------------------------------------------------- #
# 4. OPENAI CALLS - retry + on-disk cache so reruns cost nothing
# --------------------------------------------------------------------------- #
class Generator:
    def __init__(self, model: str, cache_path: Path, temperature: float = 0.4):
        self.client = OpenAI()
        self.model = model
        self.temperature = temperature
        self.cache_path = cache_path
        self.cache: dict[str, dict] = {}
        self.lock = threading.Lock()
        self.calls = 0
        self.hits = 0
        self.repairs = 0     # meta descriptions fixed by the post-checks
        self.thin_tags = 0   # rows that ended up with fewer than TAG_MIN tags
        if cache_path.exists():
            try:
                self.cache = json.loads(cache_path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                self.cache = {}

    # -- low level -------------------------------------------------------- #
    def _chat(self, system: str, user: str, json_mode: bool, max_tokens: int) -> str:
        last_error = None
        for attempt in range(5):
            try:
                kwargs = {
                    "model": self.model,
                    "messages": [
                        {"role": "system", "content": system},
                        {"role": "user", "content": user},
                    ],
                    "temperature": self.temperature,
                    "max_tokens": max_tokens,
                }
                if json_mode:
                    kwargs["response_format"] = {"type": "json_object"}
                response = self.client.chat.completions.create(**kwargs)
                with self.lock:
                    self.calls += 1
                return (response.choices[0].message.content or "").strip()
            except Exception as exc:  # rate limits, timeouts, transient 5xx
                last_error = exc
                if attempt == 4:
                    break
                time.sleep(min(2 ** attempt * 2, 30))
        raise RuntimeError(f"OpenAI request failed after 5 attempts: {last_error}")

    def _cache_key(self, stage: str, payload: str) -> str:
        raw = f"{stage}|{self.model}|{payload}"
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def _cached(self, key: str):
        with self.lock:
            if key in self.cache:
                self.hits += 1
                return self.cache[key]
        return None

    def _store(self, key: str, value) -> None:
        with self.lock:
            self.cache[key] = value

    def save_cache(self) -> None:
        with self.lock:
            self.cache_path.parent.mkdir(parents=True, exist_ok=True)
            self.cache_path.write_text(
                json.dumps(self.cache, ensure_ascii=False), encoding="utf-8"
            )

    # -- stages ----------------------------------------------------------- #
    def stage_details(self, book: dict) -> dict:
        plan = plan_translation(book)
        payload = build_details_payload(book, plan)
        key = self._cache_key("details", payload)
        cached = self._cached(key)
        if cached is None:
            raw = self._chat(DETAILS_SYSTEM_PROMPT, payload, json_mode=True, max_tokens=1200)
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                match = re.search(r"\{.*\}", raw, re.S)
                data = json.loads(match.group(0)) if match else {}

            cached = {
                "revised_name": _clean(data.get("revised_name")),
                "revised_author": _clean(data.get("revised_author")),
                "details": {
                    _clean(k): _clean(v)
                    for k, v in (data.get("details") or {}).items()
                    if _clean(v)
                },
                "description": _clean(data.get("description")),
            }
            self._store(key, cached)

        # English stays English: use the source text, never a model rewrite.
        result = dict(cached)
        if plan["name_ok"]:
            result["revised_name"] = plan["preset_name"]
        if plan["author_ok"]:
            result["revised_author"] = plan["preset_author"]
        return result

    def stage_meta(self, book: dict, stage_one: dict) -> str:
        payload = build_meta_payload(book, stage_one)
        key = self._cache_key("meta", payload)
        meta = self._cached(key)

        if meta is None:
            meta = self._chat(META_SYSTEM_PROMPT, payload, json_mode=False, max_tokens=200)
            meta = meta.strip().strip('"').strip()
            meta = re.sub(r"^meta description\s*[:\-]\s*", "", meta, flags=re.I)
            self._store(key, meta)

        # Enforced in code, not left to the prompt: the model has been seen quoting
        # an ISBN the book does not have, and overrunning the 160-char cap.
        cleaned = trim_meta(strip_wrong_isbn(meta, book.get("isbn", "")))
        if cleaned != meta:
            with self.lock:
                self.repairs += 1
        return cleaned

    def stage_tags(self, book: dict, stage_one: dict) -> str:
        payload = build_tags_payload(book, stage_one)
        key = self._cache_key("tags", payload)
        raw = self._cached(key)

        if raw is None:
            raw = self._chat(TAGS_SYSTEM_PROMPT, payload, json_mode=False, max_tokens=200)
            self._store(key, raw)

        tags = clean_tags(raw, book)
        # Under-tagging usually means the row had almost no source data. Counted
        # rather than padded: inventing a category would be worse than a short list.
        if len(tags.split(",")) < TAG_MIN:
            with self.lock:
                self.thin_tags += 1
        return tags


# --------------------------------------------------------------------------- #
# 5. HTML RENDERING - matches the layout in "image (1).png"
# --------------------------------------------------------------------------- #
# Section headings sit just 3px above the body copy - the browser default for <h3>
# is far larger and looked oversized on the product page.
BODY_FONT_PX = 16
HEADING_FONT_PX = BODY_FONT_PX + 3

HEADING_STYLE = f"font-size:{HEADING_FONT_PX}px;font-weight:700;margin:0 0 6px;"
BODY_STYLE = f"font-size:{BODY_FONT_PX}px;margin:0 0 14px;"
LIST_STYLE = f"font-size:{BODY_FONT_PX}px;margin:0 0 14px;padding-left:20px;"


def render_details_html(details: dict) -> str:
    """Ordered <ul> with bold labels; unknown extra keys are appended at the end."""
    known = [label for _, label in DETAIL_ORDER]
    ordered = [(label, details[label]) for label in known if details.get(label)]
    ordered += [(k, v) for k, v in details.items() if k not in known and v]
    items = "".join(
        f"<li><strong>{escape(label)}:</strong> {escape(value)}</li>" for label, value in ordered
    )
    return f'<ul style="{LIST_STYLE}">{items}</ul>' if items else ""


def render_listing_html(details_html: str, description: str) -> str:
    """Column G: the product body copy.

    The meta description is deliberately left out - it is taken from column F and
    belongs in the page <head>, not in the product body.
    """
    blocks = []
    if details_html:
        blocks.append(f'<h3 style="{HEADING_STYLE}">Book Details</h3>\n{details_html}')
    if description:
        blocks.append(f'<h3 style="{HEADING_STYLE}">Book Description</h3>\n'
                      f'<p style="{BODY_STYLE}">{escape(description)}</p>')
    return "\n\n".join(blocks)


def format_title(revised: str, original: str, needs_translation: bool, profile: dict) -> str:
    """Baatighar-style bilingual title: "Swapna-Bishleshon (স্বপ্ন-বিশ্লেষণ)".

    Only for sheets whose profile asks for it, and only for titles that were
    actually transliterated - an already-English title gets no bracket.
    """
    if not profile.get("bilingual_title") or not needs_translation:
        return revised
    if not revised or not original or original in revised:
        return revised
    return f"{revised} ({original})"


# --------------------------------------------------------------------------- #
# 6. WORKBOOK PROCESSING
# --------------------------------------------------------------------------- #
def process_row(generator: Generator, book: dict, profile: dict) -> dict:
    stage_one = generator.stage_details(book)

    # Both are written from the plain English title, so they are generated before
    # any bilingual bracket is added to the displayed name.
    meta = generator.stage_meta(book, stage_one)
    tags = generator.stage_tags(book, stage_one)

    description = stage_one.get("description", "")
    details_html = render_details_html(stage_one.get("details", {}))
    title = format_title(
        stage_one.get("revised_name", ""),
        book.get("name", ""),
        needs_translation=not plan_translation(book)["name_ok"],
        profile=profile,
    )
    return {
        "Revised Name": title,
        "Revised Author": stage_one.get("revised_author", ""),
        "Book Details": details_html,
        "Book Description": description,
        "Meta Description": meta,
        "Listing HTML": render_listing_html(details_html, description),
        "Tags": tags,
    }


def is_generated_file(path: Path, suffix: str) -> bool:
    """True for files this script produced.

    The name check alone is not enough - a run with a different --suffix would
    happily re-ingest a previous output - so the header row is inspected too.
    """
    if path.stem.endswith(suffix) or path.stem.endswith("_formatted"):
        return True
    try:
        workbook = openpyxl.load_workbook(path, read_only=True)
        header = {str(cell.value).strip() for cell in next(workbook.active.iter_rows(max_row=1))}
        workbook.close()
    except Exception:
        return False
    return {"Revised Name", "Book Details"}.issubset(header)


def process_workbook(path: Path, generator: Generator, args) -> Path:
    source = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = source.active

    rows = list(sheet.iter_rows(values_only=True))
    source.close()
    if not rows:
        raise ValueError(f"{path.name} is empty")

    header_row = rows[0]
    field_to_cols, leftovers = build_header_map(header_row)

    if "name" not in field_to_cols:
        raise ValueError(
            f"{path.name}: no title column found. Headers seen: "
            f"{[h for h in header_row if h]}"
        )

    log(f"\n[{path.name}] {len(rows) - 1} data rows")
    log(f"  mapped : {', '.join(sorted(field_to_cols))}")
    if leftovers:
        log(f"  extra  : {', '.join(leftovers.values())}  (passed to the model as context)")

    # Build the work list, skipping rows without a title.
    books: list[dict] = []
    for row in rows[1:]:
        book = extract_book(row, field_to_cols, leftovers, args.infer_country)
        if not book.get("name"):
            continue
        books.append(book)
        if args.limit and len(books) >= args.limit:
            break

    profile = get_profile(path, books[0].get("link", "") if books else "")
    if args.bilingual_title is not None:  # CLI override beats the sheet profile
        profile = {**profile, "bilingual_title": args.bilingual_title}
    log(f"  profile: bilingual_title={profile['bilingual_title']}")

    # VLOOKUP returns the FIRST match, so duplicate keys are silently wrong answers.
    seen: dict[str, int] = {}
    for book in books:
        seen[book["name_raw"]] = seen.get(book["name_raw"], 0) + 1
    duplicates = {key: count for key, count in seen.items() if count > 1}
    if duplicates:
        log(f"  WARNING: {len(duplicates)} duplicate product name(s) - VLOOKUP will "
              f"only ever return the first of each:")
        for key, count in list(duplicates.items())[:5]:
            log(f"      x{count}  {key[:60]}")

    if args.dry_run:
        for book in books[:3]:
            plan = plan_translation(book)
            log("\n--- payload preview ---")
            log(f"  needs transliteration: name={not plan['name_ok']} "
                  f"author={not plan['author_ok']}")
            log(build_details_payload(book, plan))
        skipped = sum(1 for b in books if plan_translation(b)["name_ok"]
                      and plan_translation(b)["author_ok"])
        log(f"\n  {skipped}/{len(books)} rows are already English "
              f"(no translation keys requested)")
        return path

    results: dict[int, dict] = {}
    failures = 0
    done = 0
    total = len(books)
    progress_lock = threading.Lock()

    def worker(job):
        nonlocal failures, done
        index, book = job
        try:
            output = process_row(generator, book, profile)
        except Exception as exc:
            with progress_lock:
                failures += 1
            log(f"  !! {book.get('name', '')[:40]}: {exc}")
            output = {column: "" for column in OUTPUT_COLUMNS}
            output["Book Details"] = f"ERROR: {exc}"
        results[index] = output
        with progress_lock:
            done += 1
            title = (output.get("Revised Name") or book.get("name", ""))[:52]
            log(f"  [{done}/{total}] {title}")

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        list(pool.map(worker, enumerate(books)))

    generator.save_cache()

    # Fresh, slim workbook: the VLOOKUP key, then any carried source columns,
    # then the generated ones. "name" is always column A, so it is never carried twice.
    carry = [field for field in args.carry
             if field != "name" and field in field_to_cols]
    headers = ([KEY_COLUMN]
               + [CARRY_HEADERS.get(field, field.title()) for field in carry]
               + OUTPUT_COLUMNS)

    out_book = openpyxl.Workbook()
    out_sheet = out_book.active
    out_sheet.title = "Books"

    for column, header in enumerate(headers, start=1):
        cell = out_sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
        out_sheet.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS.get(header, 28)
    out_sheet.freeze_panes = "A2"

    for offset, book in enumerate(books):
        excel_row = offset + 2
        output = results.get(offset, {})
        values = ([book["name_raw"]]
                  + [book.get(field, "") for field in carry]
                  + [output.get(column, "") for column in OUTPUT_COLUMNS])
        for column, value in enumerate(values, start=1):
            cell = out_sheet.cell(row=excel_row, column=column)
            cell.value = str(value or "")[:EXCEL_CELL_LIMIT]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # The key must stay a plain, untouched string - no wrapping games, and
        # never coerced to a number, or an exact-match lookup can fail.
        key_cell = out_sheet.cell(row=excel_row, column=1)
        key_cell.data_type = "s"
        key_cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_path = path.with_name(f"{path.stem}{args.suffix}.xlsx")
    out_book.save(out_path)
    log(f"  -> {out_path.name}   ({total - failures} ok, {failures} failed)")
    return out_path


# --------------------------------------------------------------------------- #
# 7. CLI
# --------------------------------------------------------------------------- #
def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate English titles/authors, book details, descriptions and "
                    "SEO meta descriptions for scraped book spreadsheets."
    )
    parser.add_argument("--folder", default=str(app_dir()),
                        help="folder containing the source .xlsx files")
    parser.add_argument("--files", nargs="*", help="specific .xlsx files (default: all in folder)")
    parser.add_argument("--suffix", default="_formatted", help="suffix for the output file")
    parser.add_argument("--model", default=os.getenv("OPENAI_MODEL", "gpt-4o-mini"))
    parser.add_argument("--workers", type=int, default=4, help="parallel API calls")
    parser.add_argument("--temperature", type=float, default=0.4)
    parser.add_argument("--limit", type=int, default=0, help="max rows per file (0 = all)")
    parser.add_argument("--infer-country", action="store_true",
                        help="fill a missing Country from the source website's domain")
    parser.add_argument("--carry", default=",".join(DEFAULT_CARRY),
                        help="optional source fields to add back into the output file, "
                             "comma separated (default: none; available: "
                             f"{', '.join(sorted(CARRY_HEADERS))})")
    parser.add_argument("--bilingual-title", dest="bilingual_title",
                        action="store_true", default=None,
                        help="force 'English (Original Script)' titles on every sheet "
                             "(default: only sheets configured for it, e.g. Baatighar)")
    parser.add_argument("--no-bilingual-title", dest="bilingual_title",
                        action="store_false",
                        help="force plain English titles on every sheet")
    parser.add_argument("--dry-run", action="store_true",
                        help="show the mapping and what would be sent, without calling the API")
    args = parser.parse_args()
    args.carry = [field.strip().lower() for field in args.carry.split(",") if field.strip()]
    return run(args)


def default_settings(**overrides):
    """A settings object with the same shape argparse produces.

    The GUI builds its run configuration through this, so both front ends share
    exactly one code path.
    """
    settings = SimpleNamespace(
        folder=str(app_dir()),
        files=None,
        suffix="_formatted",
        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
        workers=4,
        temperature=0.4,
        limit=0,
        infer_country=False,
        carry=list(DEFAULT_CARRY),
        bilingual_title=None,
        dry_run=False,
    )
    for key, value in overrides.items():
        setattr(settings, key, value)
    return settings


def app_dir() -> Path:
    """Folder to start in: next to the executable when frozen, else the source."""
    if getattr(sys, "frozen", False):
        executable = Path(sys.executable).resolve()
        # Inside a macOS .app the executable lives in Contents/MacOS, which is not
        # a useful place to look for spreadsheets - fall back to the user's home.
        if ".app/Contents/MacOS" in str(executable):
            return Path.home()
        return executable.parent
    return Path(__file__).parent


def run(args) -> int:
    """The whole batch. Shared by the CLI and the GUI."""
    if not os.getenv("OPENAI_API_KEY") and not args.dry_run:
        log("ERROR: OPENAI_API_KEY is not set. Put it in a .env file or your environment.")
        return 1

    folder = Path(args.folder)
    if args.files:
        paths = [Path(f) if Path(f).is_absolute() else folder / f for f in args.files]
    else:
        paths = sorted(
            p for p in folder.glob("*.xlsx")
            if not p.name.startswith("~$") and not is_generated_file(p, args.suffix)
        )

    if not paths:
        log(f"No source .xlsx files found in {folder}")
        return 1

    generator = None if args.dry_run else Generator(
        model=args.model,
        cache_path=folder / ".cache" / "book_formatter_cache.json",
        temperature=args.temperature,
    )
    log(f"Model: {args.model} | workers: {args.workers} | files: {len(paths)}"
          + (" | DRY RUN" if args.dry_run else ""))

    exit_code = 0
    for path in paths:
        try:
            process_workbook(path, generator, args)
        except Exception as exc:
            log(f"[{path.name}] FAILED: {exc}")
            exit_code = 1

    if generator:
        generator.save_cache()
        log(f"\nDone. API calls: {generator.calls} | cache hits: {generator.hits}"
              f" | meta descriptions repaired: {generator.repairs}"
              f" | rows with under {TAG_MIN} tags: {generator.thin_tags}")
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
